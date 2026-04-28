"""
parse_kapak_openpyxl.py
-----------------------
Reads ALL KAPAK sheets from the Excel file using openpyxl.
- Resolves merged cells (top-left value propagated to all merged cells)
- Finds every statistic row by its label text
- Extracts all numeric values using the known column layout
- Outputs parsed_kapak_data.json for inspection / database import
"""

import json
import sys
import re
from pathlib import Path
import openpyxl

# ── Column layout (0-based, matching corrected STATS_DEFAULTS in upload/route.ts) ──
# label at col 2; ROOM/PAX pairs: TODAY=3/4, MTD_ACT=5/6, MTD_BGT=7/8,
# MTD_FCST=9/10, YTD_ACT=11/12, YTD_BGT=13/14, LY_TODAY=15/16, LY_MTD=17/18, LY_YTD=19/20
SC = {
    'today': 3, 'mtd': 5, 'budget': 7, 'forecast': 9,
    'ytd': 11, 'lyToday': 15, 'lyMtd': 17, 'lyYtd': 19,
}

EXCEL_FILE = Path(__file__).parent / "Daily Report NİSAN-2026.xlsx"

# ── Merged-cell resolver ──────────────────────────────────────────────────────

def build_merged_map(ws):
    merged_map = {}
    for mr in ws.merged_cells.ranges:
        val = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_map[(r, c)] = val
    return merged_map


def cell_val(ws, merged_map, row_1, col_1):
    """Return cell value, resolving merged cells. row/col are 1-indexed (openpyxl)."""
    return merged_map.get((row_1, col_1), ws.cell(row_1, col_1).value)


def g(ws, merged_map, row_1indexed, col_0indexed):
    """Read numeric value at (row_1indexed, col_0indexed+1). Returns 0 if missing/non-numeric."""
    val = cell_val(ws, merged_map, row_1indexed, col_0indexed + 1)
    if isinstance(val, (int, float)):
        return float(val)
    return 0.0


# ── Row finder ────────────────────────────────────────────────────────────────

ROW_LABELS = {
    'availRoom':  ['AVAILABLE ROOM', 'AVAIL ROOM', 'AVAIL. ROOM'],
    'soldRoom':   ['SOLD ROOMS', 'SOLD ROOM', 'OCCUPIED ROOM'],
    'compRoom':   ['COMPS ROOMS', 'COMP ROOMS', 'COMP ROOM', 'COMP. ROOM',
                   'COMPLIMENTARY ROOM', 'COMPLIMENTARY', 'HOUSE USE',
                   'ÜCRETSİZ ODA', 'UCRETSIZ ODA', 'KOMPLİMAN', 'KOMPLIMAN'],
    'occupied':   ['OCCUPIED', 'OCCUPANCY RATE', 'OCC %', 'OCCUPANCY %'],
    'adr':        ['AVR.ROOM RATE', 'AVR. ROOM RATE', 'ADR', 'AVERAGE ROOM RATE',
                   'AVERAGE DAILY RATE', 'ORT.ODA', 'ORT. ODA'],
    'avgSales':   ['AVR.SALES RATE', 'AVR. SALES RATE', 'AVG.SALES RATE',
                   'AVERAGE SALES', 'ORT.SATIS', 'ORT. SATIŞ'],
    'payingGuest':['PAYING GUEST'],
    'compGuest':  ['COMP GUEST', 'COMPLIMENTARY GUEST'],
    'payChild':   ['PAYING CHILDREN', 'ÜCRETLİ ÇOCUK'],
    'compChild':  ['COMP CHILDREN', 'COMP CHILD'],
    'freeChild':  ['FREE CHILD', 'ÜCRETSİZ ÇOCUK'],
    'outOfOrder': ['OUT OF ORDER', 'OUT-OF-ORDER'],
    'exRateLabel':['RATE / KUR', 'EUR KURU', 'EUR/TRY', 'EURO KURU',
                   'EXCHANGE RATE', 'KUR'],
    # Revenue rows
    'totalSales': ['TOTAL SALES REVENUE', 'TOPLAM SATIŞ', 'TOPLAM SATIS'],
    'allInc':     ['ALL.INC', 'ALL INC', 'ALLINC'],
    'food':       ['F&B FOOD', 'FOOD REVENUE', 'YİYECEK', 'YIYECEK'],
    'beverage':   ['F&B BEV', 'BEVERAGE', 'İÇECEK', 'ICECEK'],
    'spa':        ['SPA'],
    'misc':       ['MISC', 'MISCELLANEOUS', 'DİĞER', 'DIGER'],
    'totals':     ['TOTALS'],
}


def find_rows(ws, merged_map):
    """Scan entire sheet and map label keys to their row numbers (1-indexed)."""
    found = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, 6):
            val = cell_val(ws, merged_map, r, c)
            if not isinstance(val, str):
                continue
            upper = val.strip().upper()
            for key, kws in ROW_LABELS.items():
                if key not in found and any(upper.startswith(kw.upper()) for kw in kws):
                    found[key] = r
    return found


# ── Value extractors ──────────────────────────────────────────────────────────

def stat_row(ws, mm, row):
    """Read a full ROOM/PAX statistics row."""
    if row < 1:
        return None
    sc = SC
    return {
        'todayRoom':     g(ws, mm, row, sc['today']),
        'todayPax':      g(ws, mm, row, sc['today'] + 1),
        'mtdRoom':       g(ws, mm, row, sc['mtd']),
        'mtdPax':        g(ws, mm, row, sc['mtd'] + 1),
        'budgetRoom':    g(ws, mm, row, sc['budget']),
        'budgetPax':     g(ws, mm, row, sc['budget'] + 1),
        'forecastRoom':  g(ws, mm, row, sc['forecast']),
        'forecastPax':   g(ws, mm, row, sc['forecast'] + 1),
        'ytdRoom':       g(ws, mm, row, sc['ytd']),
        'ytdPax':        g(ws, mm, row, sc['ytd'] + 1),
        'ytdBudgetRoom': g(ws, mm, row, sc['ytd'] + 2),
        'ytdBudgetPax':  g(ws, mm, row, sc['ytd'] + 3),
        'lyTodayRoom':   g(ws, mm, row, sc['lyToday']),
        'lyTodayPax':    g(ws, mm, row, sc['lyToday'] + 1),
        'lyMtdRoom':     g(ws, mm, row, sc['lyMtd']),
        'lyMtdPax':      g(ws, mm, row, sc['lyMtd'] + 1),
        'lyYtdRoom':     g(ws, mm, row, sc['lyYtd']),
        'lyYtdPax':      g(ws, mm, row, sc['lyYtd'] + 1),
    }


def rev_row(ws, mm, row):
    """Read a revenue row (TL/EUR pairs, then EUR-only for YTD+LY)."""
    if row < 1:
        return None
    sc = SC
    return {
        'todayTL':       g(ws, mm, row, sc['today']),
        'todayEUR':      g(ws, mm, row, sc['today'] + 1),
        'mtdActualTL':   g(ws, mm, row, sc['mtd']),
        'mtdActualEUR':  g(ws, mm, row, sc['mtd'] + 1),
        'mtdBudgetTL':   g(ws, mm, row, sc['budget']),
        'mtdBudgetEUR':  g(ws, mm, row, sc['budget'] + 1),
        'mtdForecastTL': g(ws, mm, row, sc['forecast']),
        'mtdForecastEUR':g(ws, mm, row, sc['forecast'] + 1),
        'ytdActualEUR':  g(ws, mm, row, sc['ytd']),
        'ytdBudgetEUR':  g(ws, mm, row, sc['ytd'] + 2),
        'lyTodayEUR':    g(ws, mm, row, sc['ytd'] + 4),
        'lyMonthEUR':    g(ws, mm, row, sc['ytd'] + 6),
        'lyYearEUR':     g(ws, mm, row, sc['ytd'] + 8),
    }


def exchange_rate_row(ws, mm, label_row):
    """Exchange rates are in the row BELOW the label row, at PAX (odd) columns."""
    r = label_row + 1
    sc = SC
    return {
        'dailyRate':      g(ws, mm, r, sc['today'] + 1),
        'monthlyAvgRate': g(ws, mm, r, sc['mtd'] + 1),
        'budgetRate':     g(ws, mm, r, sc['budget'] + 1),
        'forecastRate':   g(ws, mm, r, sc['forecast'] + 1),
        'yearlyAvgRate':  g(ws, mm, r, sc['ytd'] + 1),
        'yearlyBudgetRate': g(ws, mm, r, sc['ytd'] + 3),
        'lyDailyRate':    g(ws, mm, r, sc['lyToday'] + 1),
        'lyMonthlyRate':  g(ws, mm, r, sc['lyMtd'] + 1),
        'lyYearlyRate':   g(ws, mm, r, sc['lyYtd'] + 1),
    }


# ── Sheet date parser ─────────────────────────────────────────────────────────

def parse_date(sheet_name):
    name = sheet_name.replace('KAPAK', '').strip()
    m = re.match(r'^(\d{2})\.(\d{2})\.(\d{4})', name)
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    return f'{y:04d}-{mo:02d}-{d:02d}'


# ── Main parser ───────────────────────────────────────────────────────────────

def parse_kapak_sheet(ws, sheet_name):
    date_str = parse_date(sheet_name)
    if not date_str:
        return None

    mm = build_merged_map(ws)
    rows = find_rows(ws, mm)

    result = {
        'date': date_str,
        'sheet': sheet_name,
        'rowsFound': {k: v for k, v in rows.items()},
        'statistics': {},
        'revenue': {},
        'exchangeRate': None,
    }

    # Statistics rows
    for key in ['availRoom', 'soldRoom', 'compRoom', 'occupied', 'adr', 'avgSales',
                'payingGuest', 'compGuest', 'payChild', 'compChild', 'freeChild', 'outOfOrder']:
        r = rows.get(key, -1)
        if r > 0:
            result['statistics'][key] = stat_row(ws, mm, r)
        else:
            result['statistics'][key] = None

    # Revenue rows
    for key in ['totalSales', 'allInc', 'food', 'beverage', 'spa', 'misc', 'totals']:
        r = rows.get(key, -1)
        if r > 0:
            result['revenue'][key] = rev_row(ws, mm, r)
        else:
            result['revenue'][key] = None

    # Exchange rate
    ex_r = rows.get('exRateLabel', -1)
    if ex_r > 0:
        result['exchangeRate'] = exchange_rate_row(ws, mm, ex_r)

    return result


def main():
    if not EXCEL_FILE.exists():
        print(f'ERROR: Excel file not found: {EXCEL_FILE}')
        sys.exit(1)

    print(f'Loading {EXCEL_FILE.name} ...')
    wb = openpyxl.load_workbook(str(EXCEL_FILE), data_only=True)

    kapak_sheets = [s for s in wb.sheetnames if s.strip().upper().startswith('KAPAK')]
    print(f'Found {len(kapak_sheets)} KAPAK sheets: {kapak_sheets[:5]}...')

    all_data = []
    for sheet_name in kapak_sheets:
        ws = wb[sheet_name]
        parsed = parse_kapak_sheet(ws, sheet_name)
        if parsed:
            all_data.append(parsed)
            avail = parsed['statistics'].get('availRoom') or {}
            comp  = parsed['statistics'].get('compRoom')  or {}
            ex    = parsed['exchangeRate'] or {}
            print(f"  {parsed['date']}: availRoom today={avail.get('todayRoom',0):.0f}/{avail.get('todayPax',0):.0f}"
                  f"  compRoom today={comp.get('todayRoom',0):.0f}/{comp.get('todayPax',0):.0f}"
                  f"  exRate daily={ex.get('dailyRate',0):.4f}")

    out_path = Path(__file__).parent / 'parsed_kapak_data.json'
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)

    print(f'\nParsed {len(all_data)} KAPAK sheets.')
    print(f'Output saved to: {out_path}')
    print('\nThis data matches what the corrected upload/route.ts will produce.')
    print('Re-upload the Excel through the UI to populate the database.')


if __name__ == '__main__':
    main()
